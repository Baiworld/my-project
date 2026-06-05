"""FR-13 Export service — CSV/Excel generation"""

import csv
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment


def export_to_csv(rows: list[dict], headers: list[str] = None) -> str:
    output = io.StringIO()
    if not rows:
        output.write("No data\n")
        return output.getvalue()

    fieldnames = headers or list(rows[0].keys())
    writer = csv.DictWriter(output, fieldnames=fieldnames, extrasaction="ignore")
    writer.writeheader()
    writer.writerows(rows)
    return output.getvalue()


def export_to_excel(rows: list[dict], sheet_name: str = "Sheet1") -> io.BytesIO:
    output = io.BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    if not rows:
        ws["A1"] = "No data"
        wb.save(output)
        output.seek(0)
        return output

    # Header row
    headers = list(rows[0].keys())
    header_fill = PatternFill(start_color="409EFF", end_color="409EFF", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=12)

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # Data rows
    for row_idx, row in enumerate(rows, 2):
        for col_idx, header in enumerate(headers, 1):
            ws.cell(row=row_idx, column=col_idx, value=row.get(header, ""))

    # Auto-width
    for col_idx in range(1, len(headers) + 1):
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = 18

    wb.save(output)
    output.seek(0)
    return output
